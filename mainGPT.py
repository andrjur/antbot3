import json
import datetime
import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def k_sort(s):
    try:
        number = [float(match) for match in re.findall(r'\d+\.\d+', s)]
        x=number[0]
    except:
        x=3.12

    return x

def mainGPT():
    with open("result.json", 'r', encoding='utf-8') as json_file:
        a = json.load(json_file)

    names = []
    dates = []
    alldays = 800
    file_path = './tabl.xlsx'

    if os.path.exists(file_path):
        os.remove(file_path)
    else:
        print(f"Файл {file_path} не существует")

    di = {'Dates1.0': ['2001-05-10']}

    for i in range(1, len(a['messages']) ):
        x = a['messages'][i]
        thismes = [] # thismes = [419020, '2024-01-16', '14:03:57', 'Андрей 3.02 Панфёров', 'Всем привет в этом чатике!']
        for k, t in x.items():
            if k == 'id':
                thismes.append(t)
            if k == 'date':
                t = t.replace('T', ' ')
                date_time_obj = datetime.datetime.strptime(t, '%Y-%m-%d %H:%M:%S')
                if date_time_obj.date() not in dates:
                    dates.append(date_time_obj.date())
                thismes.append(str(date_time_obj.date()))
                thismes.append(str(date_time_obj.time()))
            if k == 'from':
                if t not in names:
                    names.append(t)
                thismes.append(t)
            if k == 'text':
                rrr = ''
                if isinstance(t, str):
                    rrr += t
                elif isinstance(t, list):
                    for xxx in t:
                        if isinstance(xxx, str) and len(xxx) > 3:
                            rrr += xxx
                if len(rrr) > 3:
                    rrr += '\n'
                    thismes.append(rrr)

        if len(thismes) > 4:
            name = thismes[3]
            time = thismes[2]
            text = thismes[4]

            zapis = [f'{time} {text} ']

            if thismes[1] not in di['Dates1.0']:
                di['Dates1.0'].append(thismes[1])

            row = di['Dates1.0'].index(thismes[1])

            di.setdefault(name, [' ' for _ in range(alldays)])
            di[name][row] += ' '.join(zapis)

            #if 'Юлия тройка 3.12' == name:
             #   print(name, row, '---', zapis)

    namezzz = di.keys()
    namezzz=sorted(namezzz, key=k_sort )



    while len(di['Dates1.0']) < alldays:
        di['Dates1.0'].append(' 🍦')


    tabl = pd.DataFrame(di)
    tabl=tabl[namezzz]
    tabl.to_excel(file_path)

    wb = load_workbook(file_path)
    ws = wb.active
    worksheet = wb.worksheets[0]
    worksheet.title = "Чат по участникам и дням"
    for ch in 'CDEFGHIJKLMNOPQRSTUVWXYZ':
        ws.column_dimensions[ch].width = 55
        for i in range(2,alldays):
            ws[ch+str(i)].alignment = Alignment(wrapText=True, horizontal='left', vertical='top')

    for ch in ['A', 'B', 'AA', 'AB','AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN']:
        ws.column_dimensions[ch].width = 55
        for i in range(2,alldays):
            ws[ch+str(i)].alignment = Alignment(wrapText=True, horizontal='left', vertical='top', )


    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['C'].width = 77

    ws.row_dimensions[1].height=15
    ws.row_dimensions[2].height=3

    for i in range(3,alldays+2):
        ws.row_dimensions[i].height=145

    ws.freeze_panes = "C2"

    wb.save('width55.xlsx')

    print('\n end mainGPT work \n')

if __name__ == "__main__":
    mainGPT()